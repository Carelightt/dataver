import logging
import pytz
import os 
import re 
import json # Satır numaralarını kalıcı olarak kaydetmek için

from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes, JobQueue 
from openpyxl import load_workbook

# --- Sabitler (Sana Özel Bilgiler) ---
# Lütfen bu bilgilerin doğru olduğundan emin ol.
TOKEN = "8484668521:AAGiVlPq_SAc5UKBXpC6F7weGFOShJDJ0yA"
YETKILI_USER_ID = 6672759317  # Senin Telegram Kullanıcı ID'n
HEDEF_GRUP_ID = -1003195011322 # Verilerin gönderileceği Telegram Grup ID'si
EXCEL_DOSYA_ADI = "veriler.xlsx"
KULLANILANLAR_DOSYA_ADI = "kullanilanlar.txt" # Kullanılan satırları tutacak dosya

# Veri Etiketlerinin Sıralaması (Excel sütun sırasına göre)
VERI_ETIKETLERI = [
    "TC", 
    "Ad Soyad", 
    "Telefon Numarası", 
    "Doğum Tarihi", 
    "İl/İlçe", 
    "IBAN"
]

# Günlükleme (logging) ayarları
logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)


# --- YARDIMCI KALICILIK FONKSİYONLARI ---

def kullanilan_satirlari_oku() -> set:
    """Kullanılan satır numaralarını dosyadan okur."""
    if not os.path.exists(KULLANILANLAR_DOSYA_ADI):
        return set()
    try:
        with open(KULLANILANLAR_DOSYA_ADI, 'r') as f:
            # Satır numaraları metin dosyasında JSON listesi olarak tutuluyor
            return set(json.load(f))
    except Exception as e:
        logger.warning(f"Kullanılan satırlar dosyası okunamadı, sıfırdan başlıyor: {e}")
        return set()

def kullanilan_satirlari_kaydet(satirlar: set):
    """Kullanılan satır numaralarını dosyaya kaydeder."""
    try:
        with open(KULLANILANLAR_DOSYA_ADI, 'w') as f:
            # Set nesnesini JSON'a kaydetmek için listeye çeviriyoruz
            json.dump(list(satirlar), f)
    except Exception as e:
        logger.error(f"Kullanılan satırlar kaydedilirken KRİTİK HATA: {e}")


# --- Yardımcı Fonksiyon: Yetki Kontrolü ve Yetkisiz Mesajı ---

def yetkili_mi(update: Update) -> bool:
    """
    Komutu kullanan kişinin yetkili ID'ye sahip olup olmadığını kontrol eder.
    Yetkisi yoksa istenen hata mesajını gönderir.
    """
    if update.effective_user.id != YETKILI_USER_ID:
        logger.warning(
            f"Yetkisiz erişim denemesi: User ID {update.effective_user.id} - Chat ID {update.effective_chat.id}"
        )
        update.message.reply_text("Yetkiniz yoktur.")
        return False
    return True


# --- YARDIMCI EXCEL/DURUM FONKSİYONU ---

def excel_durumu_hesapla():
    """Excel dosyasındaki kullanılan (kaydedilmiş) ve kalan (kaydedilmemiş) satırları hesaplar."""
    if not os.path.exists(EXCEL_DOSYA_ADI):
        return None, "Hata: Excel dosyası bulunamadı."
    
    try:
        workbook = load_workbook(EXCEL_DOSYA_ADI)
        sheet = workbook.active
        
        # Kullanılan satır numaralarını hafızadan oku
        kullanilan_satir_numaralari = kullanilan_satirlari_oku()
        
        kullanilan_sayisi = 0
        kalan_sayisi = 0
        baslangic_satiri = 2
        
        # Tüm veri satırlarını döngüye al
        for row_index, row in enumerate(sheet.iter_rows(min_row=baslangic_satiri), start=baslangic_satiri):
            # Satır, kullanılanlar listesinde mi kontrol et
            if row_index in kullanilan_satir_numaralari:
                kullanilan_sayisi += 1
            else:
                kalan_sayisi += 1
                
        return kalan_sayisi, kullanilan_sayisi
        
    except Exception as e:
        logger.error(f"Excel durumu hesaplanırken hata: {e}")
        return None, f"Kritik Hata: Excel okunamadı. ({e})"


# --- KOMUT İŞLEYİCİLERİ ---

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """/start komutuna yanıt verir."""
    if yetkili_mi(update):
        await update.message.reply_text(
            f'Merhaba yetkili! Ben göreve hazırım. Mevcut komutlar:\n\n'
            f'  • /ver <miktar>: Veri gönderir ve işaretler.\n'
            f'  • /kalan: Verilmemiş veri sayısını söyler.\n'
            f'  • /rapor: Verilmiş veri sayısını söyler.'
        )

async def kalan_komutu_isleyici(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Kalan veri sayısını bildirir."""
    if not yetkili_mi(update):
        return

    kalan, kullanilan = excel_durumu_hesapla()
    
    if kalan is None:
        await update.message.reply_text(kullanilan) # Hata mesajı gönderiliyor
        return
        
    await update.message.reply_text(
        f" KALAN DATA SAYISI\n"
        f"Elimizdeki data sayısı: **{kalan}**"
    )

async def rapor_komutu_isleyici(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Verilmiş (kullanılmış) veri sayısını bildirir."""
    if not yetkili_mi(update):
        return

    kalan, kullanilan = excel_durumu_hesapla()
    
    if kullanilan is None:
        await update.message.reply_text(kalan) # Hata mesajı gönderiliyor
        return
        
    await update.message.reply_text(
        f" **Gönderilen Veri Raporu**\n"
        f"Verilen data sayısı: **{kullanilan}**"
    )

async def ver_komutu_isleyici(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """
    /ver <miktar> komutunu işler. Veriyi gönderir ve satır numarasını kalıcı dosyaya kaydeder.
    Veri sayısının çok olması durumunda mesajı otomatik olarak böler ve her veriyi numaralandırır.
    """
    # 1. Yetki Kontrolü
    if not yetkili_mi(update):
        return

    # 2. Miktar Kontrolü ve Ayrıştırma
    if not context.args or not context.args[0].isdigit():
        await update.message.reply_text("Kullanım: `/ver <miktar>`. Lütfen kaç adet veri istediğinizi sayı olarak belirtin.")
        return
    
    try:
        miktar = int(context.args[0])
        if miktar <= 0:
            await update.message.reply_text("Lütfen pozitif bir sayı girin.")
            return
    except ValueError:
        await update.message.reply_text("Miktar sayı olmalıdır.")
        return

    # 3. Excel Dosya Kontrolü
    if not os.path.exists(EXCEL_DOSYA_ADI):
        await update.message.reply_text(f"Hata: '{EXCEL_DOSYA_ADI}' dosyası bulunamadı. Lütfen kontrol edin.")
        return

    await update.message.reply_text(f"Talep edilen {miktar} adet data çekiliyor ve gruba gönderiliyor...")

    try:
        # Önce mevcut kullanılan satırları oku (Kalıcılık için)
        mevcut_kullanilanlar = kullanilan_satirlari_oku()
        
        workbook = load_workbook(EXCEL_DOSYA_ADI)
        sheet = workbook.active  
        
        tum_veriler = [] # Tüm çekilen verileri tutar
        yeni_kullanilacak_satir_numaralari = [] # Bu oturumda kullanılanlar
        veri_sayisi_toplam = 0 # Toplam çekilecek veri sayısını kontrol eder
        veri_sayac = 0 # Mesaj içindeki verileri 1'den başlatarak numaralandırır
        baslangic_satiri = 2 
        
        # Okuma ve Toplama Döngüsü
        for row_index, row in enumerate(sheet.iter_rows(min_row=baslangic_satiri), start=baslangic_satiri):
            
            # TXT KONTROLÜ
            if row_index in mevcut_kullanilanlar:
                 continue

            if veri_sayisi_toplam >= miktar:
                break
                
            # Veri formatlama kısmı
            satir_verisi_duzenli = []
            hucre_degerleri = [str(cell.value).strip() if cell.value is not None else "" for cell in row]
            
            for etiket_index, etiket in enumerate(VERI_ETIKETLERI):
                if etiket_index < len(hucre_degerleri):
                    deger = hucre_degerleri[etiket_index]
                    satir_verisi_duzenli.append(f"**{etiket}**: {deger}")
            
            # Veri setinin başına numarayı ekle (YENİ)
            veri_sayac += 1
            numarali_veri = f"**{veri_sayac}. DATA**\n" + "\n".join(satir_verisi_duzenli)
            
            tum_veriler.append(numarali_veri) 
            
            yeni_kullanilacak_satir_numaralari.append(row_index)
            veri_sayisi_toplam += 1

        if not tum_veriler:
            await update.message.reply_text("Üzgünüm, Excel dosyasında gönderilebilecek işaretlenmemiş veri kalmadı.")
            return

        # 5. Verileri Gruba Bölerek Gönderme (4096 Karakter Limitini Aşmamak İçin)
        
        MAX_CHAR_LIMIT = 3800 
        VERI_AYIRICI = "\n\n---\n\n"
        
        gonderilecek_gruplar = []
        mevcut_grup = []
        mevcut_grup_uzunlugu = 0
        
        for veri in tum_veriler:
            veri_uzunlugu = len(veri) + len(VERI_AYIRICI) 
            
            if mevcut_grup_uzunlugu + veri_uzunlugu >= MAX_CHAR_LIMIT:
                gonderilecek_gruplar.append(mevcut_grup)
                mevcut_grup = []
                mevcut_grup_uzunlugu = 0
                
            mevcut_grup.append(veri)
            mevcut_grup_uzunlugu += veri_uzunlugu
            
        if mevcut_grup:
            gonderilecek_gruplar.append(mevcut_grup)

        # Tüm grupları ayrı mesajlar olarak gönder
        toplam_gonderilen_veri = 0
        for grup in gonderilecek_gruplar:
            grup_mesaji = VERI_AYIRICI.join(grup)
            
            mesaj_basligi = f"📄 **Data Paketi** ({gonderilecek_gruplar.index(grup) + 1}/{len(gonderilecek_gruplar)})\n\n"
            
            await context.bot.send_message(
                chat_id=HEDEF_GRUP_ID,
                text=mesaj_basligi + grup_mesaji,
                parse_mode='Markdown'
            )
            toplam_gonderilen_veri += len(grup)

        # 6. Kullanılan Satırları KAYDET (Kalıcılık için)
        
        mevcut_kullanilanlar.update(yeni_kullanilacak_satir_numaralari)
        kullanilan_satirlari_kaydet(mevcut_kullanilanlar)

        await update.message.reply_text(
            f"✅ İşlem Başarılı! Toplam **{toplam_gonderilen_veri}** adet data {len(gonderilecek_gruplar)} ayrı mesaj halinde gruba gönderildi ve çöp kutusuna taşındı."
        )

    except Exception as e:
        logger.error(f"Kritik hata oluştu: {e}")
        await update.message.reply_text(f"❌ Kritik Bir Hata Oluştu. Detaylar loglara kaydedildi. Hata: `{e}`")


# --- Ana Fonksiyon ---

def main() -> None:
    """Botu başlatır."""
    
    try:
        application = (
            Application.builder()
            .token(TOKEN)
            .concurrent_updates(True)
            .job_queue(JobQueue()) 
            .build()
        )
        
        # Türkiye saat dilimini ayarla
        application.job_queue.scheduler.configure(timezone=pytz.timezone('Europe/Istanbul'))

    except Exception as e:
        logger.error(f"Bot başlatma hatası: {e}")
        logger.error("Lütfen Job Queue'yu kurduğunuzdan emin olun: 'pip install \"python-telegram-bot[job-queue]\"'")
        return 

    # Komut işleyicilerini ekle
    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("ver", ver_komutu_isleyici)) 
    application.add_handler(CommandHandler("kalan", kalan_komutu_isleyici))
    application.add_handler(CommandHandler("rapor", rapor_komutu_isleyici))

    # Bot çalışmaya başlar (sürekli güncellemeleri dinler)
    logger.info("Bot başarıyla başlatıldı ve dinlemede...")
    application.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == '__main__':
    main()
